import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")


def style_worksheet(ws, headers):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    col_widths = {"pdf_name": 50, "stamp_detected_pages": 30, "north_arrow_detected_pages": 30,
                  "total_tokens": 15, "total_cost": 15, "execution_time": 18}
    for col_idx, header in enumerate(headers, 1):
        key = header.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("$", "")
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(key, 20)


def process_fine_tuned_llm():
    approach_dir = os.path.join(OUTPUT_DIR, "fine_tuned_llm")
    rows = []

    for pdf_name in sorted(os.listdir(approach_dir)):
        pdf_dir = os.path.join(approach_dir, pdf_name)
        if not os.path.isdir(pdf_dir):
            continue

        result_path = os.path.join(pdf_dir, "result.json")
        metrics_path = os.path.join(pdf_dir, "metrics.json")

        if not os.path.exists(result_path) or not os.path.exists(metrics_path):
            continue

        with open(result_path, "r") as f:
            result = json.load(f)
        with open(metrics_path, "r") as f:
            metrics = json.load(f)

        stamp_pages = []
        north_arrow_pages = []
        for i, page in enumerate(result.get("pages", []), 1):
            parsed = page.get("prediction_parsed", {})
            if parsed.get("stamp"):
                stamp_pages.append(str(i))
            if parsed.get("north_arrow"):
                north_arrow_pages.append(str(i))

        pages_list = metrics.get("pages", [])
        num_pages = len(pages_list) or 1
        total_tokens = sum(p.get("total_tokens", 0) for p in pages_list)
        total_cost = sum(p.get("total_cost_usd", 0) for p in pages_list)
        execution_time_ms = sum(p.get("estimated_time_ms", 0) for p in pages_list)
        execution_time_s = round(execution_time_ms / 1000, 2)

        rows.append({
            "pdf_name": pdf_name,
            "total_pages": num_pages,
            "stamp_detected_pages": ", ".join(stamp_pages) if stamp_pages else "None",
            "north_arrow_detected_pages": ", ".join(north_arrow_pages) if north_arrow_pages else "None",
            "total_tokens": total_tokens,
            "total_cost": round(total_cost, 6),
            "execution_time": execution_time_s,
            "avg_tokens_per_page": round(total_tokens / num_pages, 2),
            "avg_cost_per_page": round(total_cost / num_pages, 6),
            "avg_time_per_page": round(execution_time_s / num_pages, 2),
        })

    return rows


def process_standard_llm():
    approach_dir = os.path.join(OUTPUT_DIR, "standard_llm")
    rows = []

    for pdf_name in sorted(os.listdir(approach_dir)):
        pdf_dir = os.path.join(approach_dir, pdf_name)
        if not os.path.isdir(pdf_dir):
            continue

        result_path = os.path.join(pdf_dir, "result.json")
        metrics_path = os.path.join(pdf_dir, "metrics.json")

        if not os.path.exists(result_path) or not os.path.exists(metrics_path):
            continue

        with open(result_path, "r") as f:
            result = json.load(f)
        with open(metrics_path, "r") as f:
            metrics = json.load(f)

        stamp_pages = []
        north_arrow_pages = []
        for i, page in enumerate(result.get("pages", []), 1):
            stamp_res = page.get("stamp_result", {})
            if stamp_res.get("checkStampPresence", "").lower() == "yes":
                stamp_pages.append(str(i))

            na_res = page.get("north_arrow_result", {})
            if na_res.get("NorthDirectionSymbol", "").lower() == "detected":
                north_arrow_pages.append(str(i))

        pages_list = metrics.get("pages", [])
        num_pages = len(pages_list) or 1
        total_tokens = sum(p.get("total_tokens", 0) for p in pages_list)
        total_cost = sum(p.get("total_cost_usd", 0) for p in pages_list)
        execution_time_ms = sum(p.get("estimated_time_ms", 0) for p in pages_list)
        execution_time_s = round(execution_time_ms / 1000, 2)

        rows.append({
            "pdf_name": pdf_name,
            "total_pages": num_pages,
            "stamp_detected_pages": ", ".join(stamp_pages) if stamp_pages else "None",
            "north_arrow_detected_pages": ", ".join(north_arrow_pages) if north_arrow_pages else "None",
            "total_tokens": total_tokens,
            "total_cost": round(total_cost, 6),
            "execution_time": execution_time_s,
            "avg_tokens_per_page": round(total_tokens / num_pages, 2),
            "avg_cost_per_page": round(total_cost / num_pages, 6),
            "avg_time_per_page": round(execution_time_s / num_pages, 2),
        })

    return rows


def write_excel(rows, output_path, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "PDF Name",
        "Total Pages",
        "Stamp Detected Pages",
        "North Arrow Detected Pages",
        "Total Tokens",
        "Total Cost ($)",
        "Execution Time (s)",
        "Avg Tokens/Page",
        "Avg Cost/Page ($)",
        "Avg Time/Page (s)",
    ]

    ws.append(headers)
    for row_data in rows:
        ws.append([
            row_data["pdf_name"],
            row_data["total_pages"],
            row_data["stamp_detected_pages"],
            row_data["north_arrow_detected_pages"],
            row_data["total_tokens"],
            row_data["total_cost"],
            row_data["execution_time"],
            row_data["avg_tokens_per_page"],
            row_data["avg_cost_per_page"],
            row_data["avg_time_per_page"],
        ])

    style_worksheet(ws, headers)
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"Saved: {output_path} ({len(rows)} rows)")


if __name__ == "__main__":
    ft_rows = process_fine_tuned_llm()
    write_excel(ft_rows, os.path.join(OUTPUT_DIR, "fine_tuned_llm_report.xlsx"), "Fine-Tuned LLM")

    std_rows = process_standard_llm()
    write_excel(std_rows, os.path.join(OUTPUT_DIR, "standard_llm_report.xlsx"), "Standard LLM")

    print("Done!")
